#!/usr/bin/env python3
"""tbl2map: PCR tabular data mapper

Convert PCR tabular data to tube well maps.

Functions
---------
read_argv(desc='')
    Read in sys.argv.
read_yaml(yml, is_echo=False)
    Read in a YAML file encoded in UTF-8.
notify_progress(type='start', s_start='Something is in progress...',
                s_end='Completed.', border_symb='-', border_num=59,
                is_border_only=False)
    Notify that something is in progress.
expand_re(inp)
    Expand regexes into matched file names.
get_tidy_list(inps)
    Return a tidy list of file names.
concat_path_and_files(the_path, the_files)
    Concatenate a path and files.
identify_tbls(the_yml)
    Identify tabular files to be converted.
highlight_cells(val, color_map={}):
    Highlight spreadsheet cells.
set_border(ws, cell_range)
    A custom openpyxl function for setting cell borders.
convert_table_to_map(the_yml, code_name)
    Convert tabular data into tube well maps.
"""

import os
import sys
import re
import argparse
import yaml
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
import pandas as pd

__author__ = 'Jaewoong Jang (Isotope Science Center, University of Tokyo)'
__version__ = '1.0.0'
__date__ = '2022-09-02'


def read_argv(code_root,
              desc=''):
    """Read in sys.argv.

    Parameters
    ----------
    code_root : str
        The root of this Python code
    desc : str
        The description of argparse.ArgumentParser (default '')

    Returns
    -------
    argparse.Namespace
        The Namespace object of argparse
    """
    yml_subdir_default = '{}/yaml'.format(code_root)
    for d in [yml_subdir_default, '.']:
        yml_default = '{}/trial.yaml'.format(d)
        if os.path.exists(yml_default):
            break
    parser = argparse.ArgumentParser(
        description=desc,
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('--yml', '-y',
                        default=yml_default,
                        help='YAML file specifying conversion specifications')
    return parser.parse_args()


def read_yaml(yml,
              is_echo=False):
    """Read in a YAML file encoded in UTF-8.

    Parameters
    ----------
    yml : str
        YAML file to be read in
    is_echo : bool
        Dump the YAML content. (default False)

    Returns
    -------
    yaml_loaded : dict
        YAML content
    """
    if not os.path.exists(yml):
        print('YAML file not found. Terminating.')
        sys.exit()
    with open(yml, encoding='utf-8') as fh:
        yaml_loaded = yaml.load(fh, Loader=yaml.FullLoader)
    if is_echo:
        print('-' * 70)
        print('Content of [{}]'.format(yml))
        print('-' * 70)
        print(yaml.dump(yaml_loaded, sort_keys=False))
    return yaml_loaded


def notify_progress(type='start',
                    s_start='In progress...',
                    s_end='Completed.',
                    border_symb='-',
                    border_num=59,
                    is_border_only=False):
    """Notify that something is in progress.

    Parameters
    ----------
    type : str
        Notification type; start, end (default 'start')
    s_start : str
        Notification statement for type == start (default 'In progress...')
    s_end : str
        Notification statement for type == end (default 'Completed')
    border_symb : str
        Border symbol (default '-')
    border_num : int
        Number of border symbols (default 59)
    is_border_only : bool
        If True, only a line of border is printed. (default False)
    """
    border = border_symb * border_num
    if is_border_only:
        print(border)
    else:
        if re.search('(?i)start', type):
            print(border)
            print(s_start)
            print(border)
        else:
            print(s_end)


def expand_re(inps):
    """Expand regexes into matched file names.

    Parameters
    ----------
    inp : str
        A list of file names to be examined.

    Returns
    -------
    inps_expanded : str or list
        A file name or a list of file names where
        regex-matched files have been expanded.
    """
    re_hook = re.compile('(?i)!(?:re)?(?:gex)?$')
    inps_expanded = []
    for inp in inps:
        # If the filename is regex-expressed, append the matched and existing
        # files to the returning list.
        if re.search(re_hook, inp):
            # e.g. /!(?i)admin.*[.]xlsx!regex
            # _spl[0] = '/'
            # _spl[1] = '(?i)admin.*[.]xlsx'
            # _spl[2] = 'regex'
            # The regex-expressed file name itself will be discarded
            # from the returning list of file names.
            _spl = re.split('!', inp)
            inp_path = os.path.normpath(_spl[0])  # Discard the redundant sep.
            re_inp = re.compile(_spl[1])
            for f in os.listdir(inp_path):
                if re.search(re_inp, f):
                    inp_matched = '{}/{}'.format(inp_path, f)
                    inps_expanded.append(inp_matched)
        # Append nonregexed filenames verbatim to the returning list.
        else:
            inps_expanded.append(inp)
    return inps_expanded


def get_tidy_list(inps):
    """Return a tidy list of file names.

    Parameters
    ----------
    inps : list
        List of file names to be processed

    Returns
    -------
    inps : list
        List of processed file names
    """
    # Consistent path separator
    inps = [re.sub(r'[\\/]', re.escape(os.sep), inp) for inp in inps]
    # Duplicate removal
    inps = list(set(inps))
    # Sorting
    inps = sorted(inps)
    # Nonfile removal
    for inp in inps:
        inp = re.split(r'\s*;\s*', inp)[0]  # Inspect only file names.
        if not os.path.isfile(inp):
            inps.remove(inp)
    return inps


def concat_path_and_files(the_path, the_files):
    """Concatenate a path and files.

    Parameters
    ----------
    the_path : str
        The path to be concatenated
    the_files : list
        The files to be concatenated

    Returns
    -------
    files_w_path : list
        List of files joined with the path
    """
    files_w_path = []
    for _file in the_files:
        _file_w_path = '{}/{}'.format(os.path.expandvars(the_path),
                                      os.path.expandvars(_file))
        file_w_path = re.sub(r'\\', '/', _file_w_path)  # For consistency
        files_w_path.append(file_w_path)
    return files_w_path


def identify_tbls(the_yml):
    """Identify tabular files to be converted.

    Parameters
    ----------
    the_yml : dict
        YAML-generated dict containing conversion specifications

    Returns
    -------
    inps_existing : list
        List of uniqued and nontemporary tabular files found to exist
    """
    # Prepend the input path to input file names in advance.
    _inp_path = os.path.expandvars(the_yml['inp']['path'])
    _inps = the_yml['inp']['files']
    inps_to_be_expanded = concat_path_and_files(_inp_path, _inps)
    # Regex expansion and tidying
    inps = expand_re(inps_to_be_expanded)
    inps_tidy = get_tidy_list(inps)
    inps_existing = []  # Init
    for inp in inps_tidy:
        # ~$: Temporary Excel file
        if (not re.search(r'\~[$]', inp) and
                os.path.exists(inp)):
            inps_existing.append(inp)
    return inps_existing


def highlight_cells(val,
                    color_map={}):
    """Highlight spreadsheet cells.

    Parameters
    ----------
    val : str
        The value of a cell to be examined.
    color_map : dict
        Pairs of a cell value and color specifications. Actual data come from
        the YAML file if designated by the user. (default {})

    Returns
    -------
    the_color : str
        The matched color
    """
    # Init
    color_lst = []
    val_to_be_examined = str(val)
    # e.g. 'Pos Ctrl', 'Neg Ctrl'
    for val_to_be_colored in color_map:
        # e.g. val_to_be_colored: ke, val_to_be_examined: UT
        if re.search(val_to_be_colored, val_to_be_examined):
            if len(color_map[val_to_be_colored]['background_color']):
                _color = 'background-color:{}'.format(
                    color_map[val_to_be_colored]['background_color'])
                color_lst.append(_color)
            if len(color_map[val_to_be_colored]['text_color']):
                _color = 'color:{}'.format(
                    color_map[val_to_be_colored]['text_color'])
                color_lst.append(_color)
            break
    the_color = ';'.join(color_lst) if color_lst else None
    return the_color


def set_border(ws, cell_range):
    """A custom openpyxl function for setting cell borders.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
    cell_range : str
        The range of cells to be worked on
    """
    thin = Side(border_style='thin',
                color='000000')
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin,
                                 left=thin,
                                 right=thin,
                                 bottom=thin)


def convert_table_to_map(the_yml, code_name):
    """Convert tabular data into tube well maps.

    Parameters
    ----------
    the_yml : dict
        YAML-generated dict containing conversion specifications
    code_name : str
        The file name of this code
    """
    # Obtain the list of tabular files to be converted.
    tbls = identify_tbls(the_yml)
    # Obtain pairs of column header (e.g. 'Well', 'Content', 'Sample') and
    # the corresponding column numbers (e.g. 'B', 'E', 'F').
    data_cols = {}
    for pair in the_yml['inp']['data']['cols']:
        content, col = re.split(';', pair)
        data_cols[content] = col
    # Preprocessing for highlight_cells()
    color_map = {}
    if 'highlight' in the_yml['inp']['data']:
        for hlgt in the_yml['inp']['data']['highlight']:
            val_to_be_colored, bkg_clr, txt_clr = re.split(';', hlgt)
            color_map[val_to_be_colored] = {
                'background_color': bkg_clr,
                'text_color': txt_clr,
            }
    # A template DataFrame (DF) into which the map DF (converted from tabular
    # data) will be merged.
    # - Necessary to make unused tube wells visible on the resulting file.
    # - The number of wells was hardcoded here with respect to a PCR machine
    #   we have; modify them as per your needs.
    well_cols = list(range(1, 13))  # 1..12
    well_rows = [chr(i) for i in range(ord('A'), ord('H')+1)]  # A..H
    unloaded_wells = {}
    for col in well_cols:  # 1..12
        unloaded_wells[col] = [np.nan] * len(well_rows)
    df_tpl = pd.DataFrame(unloaded_wells,
                          index=well_rows)
    # Work on the designated tabular files.
    for tbl in tbls:  # tbl: [<spreadsheet_file>, <image_file>]
        # First 1: as the rows begin at 1, and the next 1: Column header
        number_of_rows_til_image = 1 + 1 + len(df_tpl.index)
        print('[{}] being processed..'.format(tbl), end='')
        # Extract the date time info.
        regex_date = re.compile('[0-9]{4}-[0-9]{2}-[0-9]{2}')
        regex_time = re.compile(' [0-9]{2}-[0-9]{2}-[0-9]{2}')
        created_dt = {'date': None, 'time': None}
        if re.search(regex_date, tbl):
            matched = re.search(regex_date, tbl).group(0)
            # e.g. 2022-09-02 to 2022/09/02
            created_dt['date'] = re.sub('-', '/', matched)
        if re.search(regex_time, tbl):
            matched = re.search(regex_time, tbl).group(0)
            # e.g. 12-26-21 to 12:26:21
            created_dt['time'] = re.sub('-', ':', matched)
        # The original DF
        # - The sheet name was hardcoded here with respect to a PCR machine
        #   we have; modify it as per your needs.
        df_orig = pd.read_excel(tbl,
                                sheet_name=the_yml['inp']['data']['ws_name'])
        df_orig.drop(['Unnamed: 0'], axis=1, inplace=True)
        # The DF to be worked on: Use only the columns of interest.
        # e.g.
        # - Well (A01, A02, ...)
        # - Content (Unkn, Pos Ctrl, Neg Ctrl)
        # - Sample (UT100, UT1001, ...)
        df = df_orig.copy()
        df = df.loc[:, data_cols.keys()]
        # Identify the locations of positive and negative controls.
        pos_neg_ctrls = df['Content'].str.contains('(?i)Ctrl')
        # Create a column called 'Label' and fill it with the positive and
        # negative controls. This will create a Pandas Series like:
        #   0 NaN
        #   1 NaN
        #   ...
        #   84 Pos Ctrl
        df['Label'] = df.loc[pos_neg_ctrls, ['Content']]
        # Replace NaN values with corresponding 'Sample' values.
        # The above Series will then become like:
        #   0 UT100
        #   1 UT1001
        #   ...
        #   84 Pos Ctrl
        df['Label'] = df['Label'].combine_first(df['Sample'])
        # Extract group identifiers from the labels (hardcoded;
        # modify it as per your needs)
        regex_group = re.compile('([a-zA-Z]+\s*[a-zA-Z]+)')
        df['Label.group'] = df['Label'].str.extract(regex_group)
        # Extract examinee identifiers from the labels (hardcoded;
        # modify it as per your needs)
        regex_examinee = re.compile('([0-9]+)')
        df['Label.examinee'] = df['Label'].str.extract(regex_examinee)
        # Split the well names into alphabets and two-digit numbers.
        # - An alphabet corresponds to the row of a well.
        # - A two-digit number corresponds to the column of a well.
        df['Well.row'] = df['Well'].str.extract('([a-zA-Z])')
        df['Well.col'] = df['Well'].str.extract('([0-9]+)').astype(np.int64)
        # A DF with new columns of interest
        cols_of_int = ['Well', 'Well.row', 'Well.col',
                       'Label', 'Label.group', 'Label.examinee']
        df = df.loc[:, cols_of_int]
        # Convert the tabular data to a tube well map
        df_pivot = df.pivot(index='Well.row',
                            columns='Well.col',
                            values='Label')
        # Merge the DF into a template DF.
        df_pivot = df_pivot.combine_first(df_tpl)
        # Highlight designated data.
        df_pivot_styled = df_pivot.style.applymap(highlight_cells,
                                                  color_map=color_map)
        # >> pandas
        # Save the converted well map to a new sheet.
        with pd.ExcelWriter(tbl,
                            mode='a',
                            if_sheet_exists='replace') as writer:
            df_pivot_styled.to_excel(writer,
                                     sheet_name=code_name)
        # <<
        # >> openpyxl
        # Add borders to the cells.
        wb = load_workbook(tbl)
        ws = wb[code_name]
        set_border(ws, 'A1:M9')
        # Insert margins around the map.
        margin_row = 2
        margin_col = 1
        number_of_rows_til_image += margin_row
        ws.insert_rows(0, amount=margin_row)
        ws.insert_cols(0, amount=margin_col)
        # Add the date information to the 'A1' cell.
        cell_date = ws['A1']
        cell_date.value = created_dt['date']
        cell_date.font = Font(bold=True)
        # Append the original tabular data.
        cols_to_be_appended = ['Well', 'Fluor', 'Target', 'Content', 'Sample',
                               'Cq', 'Cq Mean', 'Cq Std. Dev']
        df_to_be_appended = df_orig.copy()
        df_to_be_appended = df_to_be_appended.loc[:, cols_to_be_appended]
        gap_btw_map_and_tbl = 1
        number_of_rows_til_image += gap_btw_map_and_tbl
        for i in list(range(gap_btw_map_and_tbl)):
            ws.append([])
        for row in dataframe_to_rows(df_to_be_appended,
                                     index=False,
                                     header=True):
            ws.append(row)
        # Append, if any, an image.
        is_img = False
        tbl_bname = os.path.splitext(tbl)[0]
        img_exts = ['png', 'jpg', 'bmp']  # Default
        if 'img' in the_yml['inp']['data']:  # Overriding
            img_exts = the_yml['inp']['data']['img']
        for img_ext in img_exts:
            img_fname = '{}.{}'.format(tbl_bname, img_ext)
            if os.path.exists(img_fname):
                is_img = True
                break
        if is_img:
            gap_btw_tbl_and_img = 1
            img_cell_col = chr(65  # Character 'A'
                               + len(cols_to_be_appended)
                               + gap_btw_tbl_and_img)
            img_cell_loc = img_cell_col + str(number_of_rows_til_image)
            img = Image(img_fname)
            ws.add_image(img, img_cell_loc)
        # Save the workbook.
        wb.save(tbl)
        # <<
        print(' Completed.')


if __name__ == '__main__':
    the_code = sys.argv[0]  # __file__ works with .py, but NOT with .exe!
    code_name = os.path.splitext(os.path.basename(the_code))[0]
    code_root = os.path.dirname(os.path.abspath(the_code))
    argv = read_argv(code_root)
    the_yml = read_yaml(argv.yml)
    notify_progress(type='borderless',  # Any string but 'start'
                    s_end='Running {}...'.format(code_name))
    notify_progress(is_border_only=True)
    convert_table_to_map(the_yml, code_name)
    notify_progress(is_border_only=True)
    notify_progress(type='borderless',  # Any string but 'start'
                    s_end='Completed.'.format(code_name))
